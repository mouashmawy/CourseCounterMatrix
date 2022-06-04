#Some constants for your code...
first_column = 5
first_row = 2
Input_file_name = "inputExample.xlsx"
Output_file_name = "Output1.xlsx"
############################################
from openpyxl import *
from openpyxl.styles import *
from openpyxl.comments import Comment

#importing current file that contains data
wb = load_workbook(Input_file_name)
s = wb.active
#creating new file for saving info
wb2 = Workbook()
outSheet = wb2["Sheet"]
SomeDataSheet = wb2.create_sheet("SomeDataSheet", 0)
SomeDataSheet.title = "SomeDataSheet"

#Reading coyrses names from main file
courseNamesList = []
for course in range(5,s.max_column+1):
    courseNamesList.append(s.cell(row=1,column=co-urse).value)


#writing courses names to Matrix
outSheet['A1'].value ="MATRIX"
for rowandcolumn in range(len(courseNamesList)):
    outSheet.cell(row=rowandcolumn+2,column=1).value = courseNamesList[rowandcolumn]
    outSheet.cell(row=1, column=rowandcolumn + 2).value = courseNamesList[rowandcolumn]


#Reading data from main sheet and putting them in a List
AllStudentsList = []
for student in range(2,s.max_row+1):
    CoursesOfStudent = []
    for course in range(5,s.max_column+1):
        cellvalue =s.cell(row=student,column=course).value
        x = 1 if (cellvalue==1) else 0
        CoursesOfStudent.append(x)
    AllStudentsList.append(CoursesOfStudent)


#Calculating the matrix putting it in another list
AllCourseMatrix = []
for i in range(len(AllStudentsList[0])):
    OneCourseMatrix = []
    for j in range(len(AllStudentsList[0])):
        matchesCounter=0
        for k in range(len(AllStudentsList)):
            if(AllStudentsList[k][i] and AllStudentsList[k][j]):
                matchesCounter+=1
        OneCourseMatrix.append(matchesCounter)
    AllCourseMatrix.append(OneCourseMatrix)



#Assigning values from Matrix List to Matrix Sheet
for i in range(len(AllCourseMatrix)):
    for j in range(len(AllCourseMatrix[i])):
        outSheet.cell(row=i+2,column=j+2).value =  AllCourseMatrix[i][j]


#adding number of registered students in every course
SomeDataSheet['A1'].value ="Course"
SomeDataSheet['B1'].value ="Number registered"
for i in range(2,outSheet.max_row+1):
    SomeDataSheet.cell(row=i, column=1).value = outSheet.cell(row=1, column=i).value
    SomeDataSheet.cell(row=i,column=2).value = outSheet.cell(row=i,column=i).value


#color code for every number
diagonal = [0, 0, '8888FF']
first = [1,5,'88FF88']
second = [6,15,'FFFF88']
third = [16,500,'FF8888']

#for loop for every cell to make conditional formatting and comments
for i in range(2,outSheet.max_row+1):
    for j in range(2,outSheet.max_column+1):
        #conditional formatting
        outSheet.cell(row=i,column=j).alignment = Alignment(horizontal="center") #alignning it to ceter
        cellValue = outSheet.cell(row=i,column=j).value #saving cell velue

        if (cellValue>=first[0] and cellValue<=first[1]):
           outSheet.cell(row=i,column=j).fill = PatternFill(start_color=first[2],fill_type='solid')
        elif (cellValue>=second[0] and cellValue<=second[1]):
           outSheet.cell(row=i,column=j).fill = PatternFill(start_color=second[2],fill_type='solid')
        elif (cellValue>=third[0] and cellValue<=third[1]):
           outSheet.cell(row=i,column=j).fill = PatternFill(start_color=third[2],fill_type='solid')

        #comments
        course1 = outSheet.cell(row=i, column=1).value
        course2 = outSheet.cell(row=1, column=j).value
        outSheet.cell(row=i, column=j).comment = Comment(f"{course1} & {course2}","MA")

#The diagonal (same course)
for i in range(2,outSheet.max_row+1):
    outSheet.cell(row=i,column=i).fill = PatternFill(start_color=diagonal[2], fill_type='solid')

#sheet headers
for i in range(1,outSheet.max_row+1):
    outSheet.cell(row=i, column=1).font = Font(bold=True) #row
    outSheet.cell(row=i, column=1).alignment = Alignment(horizontal="center")  # alignning it to ceter
    outSheet.cell(row=1, column=i).font = Font(bold=True) #column
    outSheet.cell(row=1, column=i).alignment = Alignment(horizontal="center")  # alignning it to ceter


#saving the file
wb2._sheets = [wb2._sheets[1],wb2._sheets[0]]
wb2.save(Output_file_name)
